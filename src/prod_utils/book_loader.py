"""
Book Loader Utility

Loads new books into the database from an Excel file.
Reads book metadata from Excel and extracts PDF metadata (size and page count).

Dependencies:
    pip install psycopg2-binary python-dotenv openpyxl PyMuPDF

Usage:
    python book_loader.py
"""

import os
from pathlib import Path
from typing import Optional, Dict, Any, List
import logging
import fitz  # PyMuPDF
import openpyxl
from pure_bhakti_vault_db import PureBhaktiVaultDB, DatabaseError


class BookLoader:
    """
    Utility to load books from Excel file into PostgreSQL database.
    """

    def __init__(
        self,
        excel_path: str,
        pdf_folder: str,
        db: Optional[PureBhaktiVaultDB] = None
    ):
        """
        Initialize the book loader.

        Args:
            excel_path: Path to the Excel file with book data
            pdf_folder: Path to the folder containing PDF files
            db: Optional PureBhaktiVaultDB instance
        """
        self.excel_path = Path(excel_path)
        self.pdf_folder = Path(pdf_folder)
        self.db = db or PureBhaktiVaultDB()
        self.logger = self._setup_logger()

        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        if not self.pdf_folder.exists():
            raise FileNotFoundError(f"PDF folder not found: {pdf_folder}")

    def _setup_logger(self) -> logging.Logger:
        """Setup logging for the book loader."""
        logger = logging.getLogger(__name__)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger

    def _get_pdf_metadata(self, pdf_name: str) -> Dict[str, Any]:
        """
        Extract metadata from PDF file.

        Args:
            pdf_name: Name of the PDF file

        Returns:
            dict: Dictionary with 'file_size_bytes' and 'number_of_pages'
        """
        pdf_path = self.pdf_folder / pdf_name
        metadata = {
            'file_size_bytes': None,
            'number_of_pages': None
        }

        if not pdf_path.exists():
            self.logger.warning(f"PDF file not found: {pdf_path}")
            return metadata

        try:
            # Get file size
            metadata['file_size_bytes'] = pdf_path.stat().st_size

            # Get page count using PyMuPDF
            doc = fitz.open(str(pdf_path))
            metadata['number_of_pages'] = len(doc)
            doc.close()

            self.logger.info(
                f"PDF {pdf_name}: {metadata['number_of_pages']} pages, "
                f"{metadata['file_size_bytes']:,} bytes"
            )

        except Exception as e:
            self.logger.error(f"Error reading PDF {pdf_name}: {e}")

        return metadata

    def _parse_int4range(self, value: Any) -> Optional[str]:
        """
        Parse and format int4range values for PostgreSQL.

        Args:
            value: Range value (e.g., '[3,7)', '3-7', or None)

        Returns:
            str: Formatted PostgreSQL int4range string or None
        """
        if value is None or value == '':
            return None

        value_str = str(value).strip()

        # Already in PostgreSQL format
        if value_str.startswith('[') or value_str.startswith('('):
            return value_str

        # Handle simple formats like '3-7' or '3,7'
        if '-' in value_str or ',' in value_str:
            parts = value_str.replace('-', ',').split(',')
            if len(parts) == 2:
                try:
                    start = int(parts[0].strip())
                    end = int(parts[1].strip())
                    # Use exclusive upper bound (standard PostgreSQL int4range)
                    return f'[{start},{end + 1})'
                except ValueError:
                    self.logger.warning(f"Could not parse range: {value_str}")
                    return None

        return value_str

    def _read_excel_data(self) -> List[Dict[str, Any]]:
        """
        Read book data from Excel file.

        Returns:
            list: List of dictionaries with book data
        """
        self.logger.info(f"Reading Excel file: {self.excel_path}")

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active

        # Get column names from first row
        headers = [cell.value for cell in ws[1]]
        self.logger.info(f"Excel columns: {headers}")

        books = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue

            # Create dictionary from headers and row values
            book_data = dict(zip(headers, row))

            # Skip if no PDF name
            if not book_data.get('pdf_name'):
                self.logger.warning(f"Row {row_idx}: No pdf_name, skipping")
                continue

            books.append(book_data)

        self.logger.info(f"Read {len(books)} books from Excel")
        return books

    def _insert_book(self, book_data: Dict[str, Any]) -> Optional[int]:
        """
        Insert a book into the database.

        Args:
            book_data: Dictionary with book data

        Returns:
            int: book_id of inserted record, or None if failed
        """
        pdf_name = book_data.get('pdf_name')

        # Check if book already exists
        existing_id = self.db.get_book_id_by_pdf_name(pdf_name)
        if existing_id:
            self.logger.warning(
                f"Book '{pdf_name}' already exists with ID {existing_id}, skipping"
            )
            return None

        # Get PDF metadata if file exists
        pdf_metadata = self._get_pdf_metadata(pdf_name)

        # Prepare insert data - use PDF metadata if Excel has None
        insert_data = {
            'pdf_name': pdf_name,
            'original_book_title': book_data.get('original_book_title'),
            'english_book_title': book_data.get('english_book_title'),
            'edition': book_data.get('edition'),
            'number_of_pages': book_data.get('number_of_pages') or pdf_metadata['number_of_pages'],
            'file_size_bytes': book_data.get('file_size_bytes') or pdf_metadata['file_size_bytes'],
            'original_author': book_data.get('original_author'),
            'commentary_author': book_data.get('commentary_author'),
            'header_height': book_data.get('header_height'),
            'footer_height': book_data.get('footer_height'),
            'page_label_location': book_data.get('page_label_location'),
            'toc_pages': self._parse_int4range(book_data.get('toc_pages')),
            'verse_pages': self._parse_int4range(book_data.get('verse_pages')),
            'glossary_pages': self._parse_int4range(book_data.get('glossary_pages')),
            'book_summary': book_data.get('book_summary')
        }

        # Build INSERT query
        columns = [k for k, v in insert_data.items() if v is not None]
        values = [insert_data[k] for k in columns]
        placeholders = ', '.join(['%s'] * len(columns))
        column_names = ', '.join(columns)

        query = f"""
            INSERT INTO book ({column_names})
            VALUES ({placeholders})
            RETURNING book_id
        """

        try:
            with self.db.get_cursor() as cursor:
                cursor.execute(query, values)
                result = cursor.fetchone()
                book_id = result['book_id'] if result else None

                if book_id:
                    self.logger.info(
                        f"✓ Inserted book '{pdf_name}' with ID {book_id}"
                    )
                return book_id

        except Exception as e:
            self.logger.error(f"✗ Error inserting book '{pdf_name}': {e}")
            raise DatabaseError(f"Failed to insert book: {e}")

    def load_books(self, dry_run: bool = False) -> Dict[str, int]:
        """
        Load all books from Excel into database.

        Args:
            dry_run: If True, only validate data without inserting

        Returns:
            dict: Statistics with 'inserted', 'skipped', 'errors'
        """
        stats = {'inserted': 0, 'skipped': 0, 'errors': 0}

        books = self._read_excel_data()

        self.logger.info(f"{'DRY RUN: ' if dry_run else ''}Processing {len(books)} books...")

        for idx, book_data in enumerate(books, start=1):
            pdf_name = book_data.get('pdf_name')
            self.logger.info(f"\n[{idx}/{len(books)}] Processing: {pdf_name}")

            try:
                if dry_run:
                    # Just validate PDF exists and can be read
                    pdf_metadata = self._get_pdf_metadata(pdf_name)
                    if pdf_metadata['number_of_pages']:
                        self.logger.info(f"  ✓ Would insert: {pdf_name}")
                        stats['inserted'] += 1
                    else:
                        self.logger.warning(f"  ✗ PDF not found: {pdf_name}")
                        stats['errors'] += 1
                else:
                    # Actually insert the book
                    book_id = self._insert_book(book_data)
                    if book_id:
                        stats['inserted'] += 1
                    else:
                        stats['skipped'] += 1

            except Exception as e:
                self.logger.error(f"  ✗ Error processing {pdf_name}: {e}")
                stats['errors'] += 1

        # Print summary
        self.logger.info("\n" + "=" * 60)
        self.logger.info("SUMMARY:")
        self.logger.info(f"  Inserted: {stats['inserted']}")
        self.logger.info(f"  Skipped:  {stats['skipped']}")
        self.logger.info(f"  Errors:   {stats['errors']}")
        self.logger.info("=" * 60)

        return stats


def main():
    """Main function to run the book loader."""

    # Configuration
    EXCEL_PATH = "/Users/kamaldivi/Development/pbb_books/tobe_processed/harmonist_book_loader.xlsx"
    PDF_FOLDER = "/Users/kamaldivi/Development/pbb_books/Harmonist/unsec"

    # Initialize loader
    loader = BookLoader(
        excel_path=EXCEL_PATH,
        pdf_folder=PDF_FOLDER
    )

    # Test database connection
    if not loader.db.test_connection():
        print("❌ Failed to connect to database. Check your .env file.")
        return

    # Run in dry-run mode first to validate
    print("\n🔍 Running in DRY RUN mode to validate data...\n")
    dry_run_stats = loader.load_books(dry_run=True)

    # Ask user to confirm
    if dry_run_stats['errors'] > 0:
        print(f"\n⚠️  Found {dry_run_stats['errors']} errors during validation.")
        response = input("Do you want to proceed anyway? (yes/no): ")
        if response.lower() != 'yes':
            print("Aborted.")
            return

    response = input("\n✅ Proceed with actual insertion? (yes/no): ")
    if response.lower() == 'yes':
        print("\n📚 Loading books into database...\n")
        loader.load_books(dry_run=False)
    else:
        print("Aborted.")


if __name__ == "__main__":
    main()
